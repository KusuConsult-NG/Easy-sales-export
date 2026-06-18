import re
import os
import csv
from openpyxl import Workbook

# Regex for illegal XML control characters (ASCII 0-8, 11, 12, 14-31)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def sanitize_value(val):
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub('', val)
    return val

def convert_csv_to_xlsx(csv_path, xlsx_path):
    print(f"Converting {csv_path} to {xlsx_path}...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, val in enumerate(row):
                val = sanitize_value(val)
                # Try to write numbers as numeric types, otherwise fallback to string
                try:
                    if val.isdigit():
                        ws.cell(row=r+1, column=c+1, value=int(val))
                    else:
                        try:
                            ws.cell(row=r+1, column=c+1, value=float(val))
                        except ValueError:
                            ws.cell(row=r+1, column=c+1, value=val)
                except Exception:
                    ws.cell(row=r+1, column=c+1, value=val)
                    
    wb.save(xlsx_path)

def main():
    exports_dir = "exports"
    
    # 1. Convert summary
    summary_csv = os.path.join(exports_dir, "users_per_state_summary.csv")
    summary_xlsx = os.path.join(exports_dir, "users_per_state_summary.xlsx")
    if os.path.exists(summary_csv):
        convert_csv_to_xlsx(summary_csv, summary_xlsx)
        os.remove(summary_csv)
        
    # 2. Convert master
    master_csv = os.path.join(exports_dir, "all_users.csv")
    master_xlsx = os.path.join(exports_dir, "all_users.xlsx")
    if os.path.exists(master_csv):
        convert_csv_to_xlsx(master_csv, master_xlsx)
        os.remove(master_csv)
        
    # 3. Convert state files
    states_dir = os.path.join(exports_dir, "users_by_state")
    if os.path.exists(states_dir):
        for file in os.listdir(states_dir):
            if file.endswith(".csv"):
                csv_path = os.path.join(states_dir, file)
                xlsx_name = file[:-4] + ".xlsx"
                xlsx_path = os.path.join(states_dir, xlsx_name)
                convert_csv_to_xlsx(csv_path, xlsx_path)
                os.remove(csv_path)
                
    print("Conversion complete!")

if __name__ == "__main__":
    main()
